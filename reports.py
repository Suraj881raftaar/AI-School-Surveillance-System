"""
reports.py
AI School Surveillance System
"""

from __future__ import annotations

import csv
import os
import sqlite3
import tkinter as tk
import threading
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, ttk

from config import *


class ReportsFrame(ttk.Frame):

    def __init__(self, parent, app=None):
        super().__init__(parent)
        self.app = app

        # Filter Variables
        self.search_var = tk.StringVar()
        self.date_range_var = tk.StringVar(value="All")
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()
        self.type_filter_var = tk.StringVar(value="All")
        self.status_filter_var = tk.StringVar(value="All")
        self.status_text = tk.StringVar(value="Ready")

        self.attendance_records: list[tuple] = []
        self.alert_records: list[tuple] = []

        self.configure(style="Reports.TFrame")
        self.configure_styles()
        self.create_widgets()
        self.load_filtered_reports()

    def configure_styles(self):
        style = ttk.Style()
        style.configure("Reports.TFrame", background=BG_COLOR)
        style.configure("ReportsPanel.TFrame", background=WHITE)
        style.configure("Reports.TLabelframe", background=BG_COLOR)
        style.configure("Reports.TLabelframe.Label", background=BG_COLOR, foreground=HEADER_COLOR, font=("Arial", 11, "bold"))
        
        style.configure(
            "Primary.TButton",
            background=BUTTON_COLOR,
            foreground=WHITE,
            font=BUTTON_FONT,
            padding=(10, 7),
        )
        style.map(
            "Primary.TButton",
            background=[("active", BUTTON_HOVER), ("pressed", BUTTON_HOVER)],
            foreground=[("active", WHITE)],
        )
        style.configure(
            "Success.TButton",
            background=SUCCESS_COLOR,
            foreground=WHITE,
            font=BUTTON_FONT,
            padding=(10, 7),
        )
        style.map(
            "Success.TButton",
            background=[("active", "#256D2B"), ("pressed", "#1E5A23")],
            foreground=[("active", WHITE)],
        )

    def create_widgets(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # -----------------------------
        # Left Side Filters Panel (Col 0)
        # -----------------------------
        filters_panel = ttk.Frame(self, style="ReportsPanel.TFrame", padding=15, relief="solid")
        filters_panel.grid(row=0, column=0, sticky="ns", padx=(0, 15), pady=15)
        filters_panel.grid_columnconfigure(0, weight=1)

        tk.Label(
            filters_panel, 
            text="Filter Controls", 
            bg=WHITE, 
            fg=HEADER_COLOR, 
            font=("Arial", 14, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 12))

        # Search Bar
        self.add_filter_label(filters_panel, "Keyword Search:", 1)
        ttk.Entry(filters_panel, textvariable=self.search_var, width=24).grid(row=2, column=0, sticky="w", pady=(0, 8))

        # Date range Selection
        self.add_filter_label(filters_panel, "Date Range Selection:", 3)
        self.date_combo = ttk.Combobox(
            filters_panel, 
            textvariable=self.date_range_var, 
            values=["All", "Today", "Yesterday", "Last 7 Days", "Last 30 Days", "Custom Range"], 
            state="readonly", 
            width=22
        )
        self.date_combo.grid(row=4, column=0, sticky="w", pady=(0, 8))
        self.date_combo.bind("<<ComboboxSelected>>", self.toggle_date_entries)

        # Custom Start & End Date Inputs (managed by dynamic display)
        self.start_lbl = tk.Label(filters_panel, text="Start Date (YYYY-MM-DD):", bg=WHITE, font=("Arial", 9, "bold"))
        self.start_entry = ttk.Entry(filters_panel, textvariable=self.start_date_var, width=24)
        self.end_lbl = tk.Label(filters_panel, text="End Date (YYYY-MM-DD):", bg=WHITE, font=("Arial", 9, "bold"))
        self.end_entry = ttk.Entry(filters_panel, textvariable=self.end_date_var, width=24)

        # Type Select
        self.add_filter_label(filters_panel, "Person Category Filter:", 9)
        self.type_combo = ttk.Combobox(
            filters_panel, 
            textvariable=self.type_filter_var, 
            values=["All", "Student", "Teacher"], 
            state="readonly", 
            width=22
        )
        self.type_combo.grid(row=10, column=0, sticky="w", pady=(0, 8))

        # Status select
        self.add_filter_label(filters_panel, "Presence Status Filter:", 11)
        self.status_combo = ttk.Combobox(
            filters_panel, 
            textvariable=self.status_filter_var, 
            values=["All", "Present", "Late"], 
            state="readonly", 
            width=22
        )
        self.status_combo.grid(row=12, column=0, sticky="w", pady=(0, 15))

        # Apply Filters Button
        ttk.Button(filters_panel, text="Apply Filter Settings", command=self.load_filtered_reports, style="Primary.TButton").grid(row=13, column=0, sticky="ew", pady=5)
        ttk.Button(filters_panel, text="Reset Filters", command=self.reset_filters).grid(row=14, column=0, sticky="ew", pady=5)

        # -----------------------------
        # Right Side Content notebooks (Col 1)
        # -----------------------------
        right_panel = ttk.Frame(self, style="Reports.TFrame")
        right_panel.grid(row=0, column=1, sticky="nsew", pady=15, padx=(0, 15))
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(1, weight=1)

        # Toolbar
        toolbar = ttk.Frame(right_panel, style="Reports.TFrame")
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        tk.Label(toolbar, text="REPORTS & ANALYTICS VIEWER", bg=BG_COLOR, fg=HEADER_COLOR, font=("Arial", 16, "bold")).pack(side="left")

        # Exporters
        ttk.Button(toolbar, text="Export PDF", command=self.export_pdf, style="Danger.TButton").pack(side="right", padx=5)
        ttk.Button(toolbar, text="Export Excel", command=self.export_excel, style="Success.TButton").pack(side="right", padx=5)
        ttk.Button(toolbar, text="Export CSV", command=self.export_csv, style="Primary.TButton").pack(side="right", padx=5)

        # Tab views
        notebook = ttk.Notebook(right_panel)
        notebook.grid(row=1, column=0, sticky="nsew")

        # Tab 1: Attendance
        attendance_frame = ttk.Frame(notebook, padding=10)
        notebook.add(attendance_frame, text="Attendance Log Summary")
        attendance_frame.grid_columnconfigure(0, weight=1)
        attendance_frame.grid_rowconfigure(0, weight=1)

        att_cols = ("ID", "Name", "Type", "ID Code", "Date", "Time", "Confidence", "Camera", "Method", "Status")
        self.attendance_table = ttk.Treeview(attendance_frame, columns=att_cols, show="headings")
        
        for col in att_cols:
            self.attendance_table.heading(col, text=col)
            self.attendance_table.column(col, width=75, anchor="center")
        self.attendance_table.column("Name", width=140, anchor="w")

        att_scroll = ttk.Scrollbar(attendance_frame, orient="vertical", command=self.attendance_table.yview)
        self.attendance_table.configure(yscrollcommand=att_scroll.set)
        self.attendance_table.grid(row=0, column=0, sticky="nsew")
        att_scroll.grid(row=0, column=1, sticky="ns")

        # Tab 2: Alerts
        alerts_frame = ttk.Frame(notebook, padding=10)
        notebook.add(alerts_frame, text="Security Alerts Log")
        alerts_frame.grid_columnconfigure(0, weight=1)
        alerts_frame.grid_rowconfigure(0, weight=1)

        alert_cols = ("ID", "Alert Type", "Description", "Date", "Time")
        self.alert_table = ttk.Treeview(alerts_frame, columns=alert_cols, show="headings")

        for col in alert_cols:
            self.alert_table.heading(col, text=col)
            self.alert_table.column(col, width=120, anchor="center")
        self.alert_table.column("Description", width=220, anchor="w")

        alert_scroll = ttk.Scrollbar(alerts_frame, orient="vertical", command=self.alert_table.yview)
        self.alert_table.configure(yscrollcommand=alert_scroll.set)
        self.alert_table.grid(row=0, column=0, sticky="nsew")
        alert_scroll.grid(row=0, column=1, sticky="ns")

        # Status Bar
        status_bar = ttk.Frame(right_panel, style="Reports.TFrame")
        status_bar.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        tk.Label(status_bar, textvariable=self.status_text, bg=BG_COLOR, fg="#555555", font=("Arial", 9)).pack(side="left")

    def add_filter_label(self, parent, text, row):
        tk.Label(parent, text=text, bg=WHITE, font=("Arial", 10, "bold"), fg="#555555").grid(row=row, column=0, sticky="w", pady=(5, 3))

    def toggle_date_entries(self, event=None):
        if self.date_range_var.get() == "Custom Range":
            self.start_lbl.grid(row=5, column=0, sticky="w", pady=(2, 2))
            self.start_entry.grid(row=6, column=0, sticky="w", pady=(0, 6))
            self.end_lbl.grid(row=7, column=0, sticky="w", pady=(2, 2))
            self.end_entry.grid(row=8, column=0, sticky="w", pady=(0, 6))
        else:
            self.start_lbl.grid_forget()
            self.start_entry.grid_forget()
            self.end_lbl.grid_forget()
            self.end_entry.grid_forget()

    def reset_filters(self):
        self.search_var.set("")
        self.date_range_var.set("All")
        self.type_filter_var.set("All")
        self.status_filter_var.set("All")
        self.toggle_date_entries()
        self.load_filtered_reports()

    def calculate_dates(self) -> tuple[str | None, str | None]:
        rng = self.date_range_var.get()
        today = datetime.now()
        
        if rng == "Today":
            d = today.strftime("%Y-%m-%d")
            return d, d
        elif rng == "Yesterday":
            d = (today - timedelta(days=1)).strftime("%Y-%m-%d")
            return d, d
        elif rng == "Last 7 Days":
            d_start = (today - timedelta(days=6)).strftime("%Y-%m-%d")
            return d_start, today.strftime("%Y-%m-%d")
        elif rng == "Last 30 Days":
            d_start = (today - timedelta(days=29)).strftime("%Y-%m-%d")
            return d_start, today.strftime("%Y-%m-%d")
        elif rng == "Custom Range":
            return self.start_date_var.get().strip() or None, self.end_date_var.get().strip() or None
            
        return None, None

    def load_filtered_reports(self):
        # Clear tables
        for child in self.attendance_table.get_children():
            self.attendance_table.delete(child)
        for child in self.alert_table.get_children():
            self.alert_table.delete(child)

        start_date, end_date = self.calculate_dates()
        keyword = self.search_var.get().strip()
        person_type = self.type_filter_var.get()
        status = self.status_filter_var.get()

        conn = None
        try:
            conn = sqlite3.connect(DATABASE_PATH)
            cur = conn.cursor()

            # --- 1. Filter Attendance ---
            att_query = """
                SELECT 
                    a.id, a.person_name, a.person_type, 
                    COALESCE(s.roll_no, t.employee_id, 'N/A') AS identity_code,
                    a.date, a.time, a.confidence, a.camera_id, a.recognition_method, a.status
                FROM attendance a
                LEFT JOIN students s ON a.person_type = 'Student' AND a.person_id = s.id
                LEFT JOIN teachers t ON a.person_type = 'Teacher' AND a.person_id = t.id
                WHERE 1=1
            """
            att_params = []

            if start_date and end_date:
                att_query += " AND a.date BETWEEN ? AND ?"
                att_params.extend([start_date, end_date])
            elif start_date:
                att_query += " AND a.date = ?"
                att_params.append(start_date)

            if person_type != "All":
                att_query += " AND a.person_type = ?"
                att_params.append(person_type)

            if status != "All":
                att_query += " AND a.status = ?"
                att_params.append(status)

            if keyword:
                att_query += " AND (a.person_name LIKE ? OR identity_code LIKE ?)"
                att_params.extend([f"%{keyword}%", f"%{keyword}%"])

            att_query += " ORDER BY a.id DESC"
            cur.execute(att_query, att_params)
            self.attendance_records = cur.fetchall()

            for r in self.attendance_records:
                # Format confidence value beautifully
                conf_val = f"{r[6]:.1f}" if r[6] is not None else "N/A"
                disp_row = (r[0], r[1], r[2], r[3], r[4], r[5], conf_val, r[7] if r[7] is not None else "N/A", r[8] or "N/A", r[9])
                self.attendance_table.insert("", "end", values=disp_row)

            # --- 2. Filter Alerts ---
            al_query = "SELECT id, alert_type, description, date, time FROM alerts WHERE 1=1"
            al_params = []

            if start_date and end_date:
                al_query += " AND date BETWEEN ? AND ?"
                al_params.extend([start_date, end_date])
            elif start_date:
                al_query += " AND date = ?"
                al_params.append(start_date)

            if keyword:
                al_query += " AND (alert_type LIKE ? OR description LIKE ?)"
                al_params.extend([f"%{keyword}%", f"%{keyword}%"])

            al_query += " ORDER BY id DESC"
            cur.execute(al_query, al_params)
            self.alert_records = cur.fetchall()

            for r in self.alert_records:
                self.alert_table.insert("", "end", values=r)

            self.status_text.set(f"Loaded {len(self.attendance_records)} attendance logs and {len(self.alert_records)} security alerts.")

        except Exception as e:
            messagebox.showerror("Database Query Error", str(e))
        finally:
            if conn:
                conn.close()

    def export_csv(self):
        if not self.attendance_records:
            messagebox.showwarning("Export CSV", "No attendance records found to export.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV File", "*.csv")],
            title="Save Attendance CSV"
        )
        if not filename:
            return

        try:
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["ID", "Name", "Type", "ID Code", "Date", "Time", "Confidence", "Camera Index", "Method", "Status"])
                writer.writerows(self.attendance_records)
            messagebox.showinfo("Success", "CSV exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def export_excel(self):
        if not self.attendance_records:
            messagebox.showwarning("Export Excel", "No attendance records found to export.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")],
            title="Save Attendance Excel"
        )
        if not filename:
            return

        # Run Excel export in background daemon thread
        threading.Thread(
            target=self.bg_export_excel,
            args=(filename,),
            daemon=True
        ).start()

    def bg_export_excel(self, filename: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messagebox.showerror("Missing Library", "openpyxl is required to export Excel files.\n\npip install openpyxl")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Surveillance Report"
            ws.views.sheetView[0].showGridLines = True

            # Headers
            headers = ["Record ID", "Name", "Type", "ID Code", "Date", "Time", "Confidence Distance", "Camera Index", "Method", "Status"]
            ws.append(headers)

            # Styling
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            for col in range(1, 11):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            row_font = Font(name="Segoe UI", size=10)
            center_align = Alignment(horizontal="center")
            left_align = Alignment(horizontal="left")

            for r in self.attendance_records:
                row_data = [
                    r[0], r[1], r[2], r[3], r[4], r[5],
                    f"{r[6]:.1f}" if r[6] is not None else "N/A",
                    r[7] if r[7] is not None else "N/A",
                    r[8] or "N/A", r[9]
                ]
                ws.append(row_data)

                curr_row = ws.max_row
                ws.cell(row=curr_row, column=1).alignment = center_align
                ws.cell(row=curr_row, column=2).alignment = left_align
                ws.cell(row=curr_row, column=3).alignment = center_align
                ws.cell(row=curr_row, column=4).alignment = center_align
                ws.cell(row=curr_row, column=5).alignment = center_align
                ws.cell(row=curr_row, column=6).alignment = center_align
                ws.cell(row=curr_row, column=7).alignment = center_align
                ws.cell(row=curr_row, column=8).alignment = center_align
                ws.cell(row=curr_row, column=9).alignment = center_align
                ws.cell(row=curr_row, column=10).alignment = center_align

                for col in range(1, 11):
                    ws.cell(row=curr_row, column=col).font = row_font

            # Fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

            wb.save(filename)
            messagebox.showinfo("Success", "Excel file exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def export_pdf(self) -> None:
        if not self.attendance_records and not self.alert_records:
            messagebox.showwarning("Export PDF", "No records found to export.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF File", "*.pdf")],
            title="Save PDF Report"
        )
        if not filename:
            return

        threading.Thread(
            target=self.bg_export_pdf,
            args=(filename,),
            daemon=True
        ).start()

    def bg_export_pdf(self, filename: str) -> None:
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "reportlab is required to export PDF files.\n\nPlease install it using:\npip install reportlab"
            )
            return

        try:
            doc = SimpleDocTemplate(filename, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
            styles = getSampleStyleSheet()

            # Custom styles
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=20,
                textColor=colors.HexColor('#0B3D91'),
                spaceAfter=15
            )
            meta_style = ParagraphStyle(
                'ReportMeta',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=9,
                textColor=colors.HexColor('#555555'),
                spaceAfter=10
            )

            story = []

            story.append(Paragraph("AI School Surveillance System", title_style))
            story.append(Paragraph("ATTENDANCE AND ALERTS REPORT SUMMARY", styles['Heading2']))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
            story.append(Paragraph("School Surveillance Engine - Analytics & Reporting", meta_style))
            story.append(Spacer(1, 15))

            if self.attendance_records:
                story.append(Paragraph("Attendance Log Records", styles['Heading3']))
                story.append(Spacer(1, 8))

                table_data = [["ID", "Name", "Type", "ID Code", "Date", "Time", "Status"]]
                # Limit rows to fit printable document size cleanly
                for r in self.attendance_records[:60]:
                    table_data.append([
                        str(r[0]),
                        str(r[1][:22]),
                        str(r[2]),
                        str(r[3]),
                        str(r[4]),
                        str(r[5]),
                        str(r[9])
                    ])

                t = Table(table_data, colWidths=[35, 160, 65, 75, 75, 65, 65])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0B3D91')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('ALIGN', (1,1), (1,-1), 'LEFT'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 9),
                    ('BOTTOMPADDING', (0,0), (-1,0), 6),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F5F7FA')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
                    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                    ('FONTSIZE', (0,1), (-1,-1), 8),
                ]))
                story.append(t)

            if self.alert_records:
                story.append(Spacer(1, 20))
                story.append(Paragraph("Security Alerts Log Records", styles['Heading3']))
                story.append(Spacer(1, 8))

                alert_data = [["ID", "Type", "Description", "Date", "Time"]]
                for r in self.alert_records[:30]:
                    alert_data.append([
                        str(r[0]),
                        str(r[1]),
                        str(r[2][:35]),
                        str(r[3]),
                        str(r[4])
                    ])

                t_alert = Table(alert_data, colWidths=[40, 110, 220, 90, 80])
                t_alert.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C62828')),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('ALIGN', (2,1), (2,-1), 'LEFT'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 9),
                    ('BOTTOMPADDING', (0,0), (-1,0), 6),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F5F7FA')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
                    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                    ('FONTSIZE', (0,1), (-1,-1), 8),
                ]))
                story.append(t_alert)

            doc.build(story)
            messagebox.showinfo("Success", "PDF report file generated successfully.")
        except Exception as e:
            messagebox.showerror("Export PDF Error", str(e))
